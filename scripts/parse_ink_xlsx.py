#!/usr/bin/env python3
"""Parse Ink Homes CRM Aug.xlsx → normalized JSON for /api/setup/migrate-data."""
import json, sys, datetime
import openpyxl

SRC = "/opt/data/profiles/rook-hw-team/cache/documents/doc_65c88a9c8529_Ink Homes CRM Aug.xlsx"
OUT = "/tmp/ink_migration"
import os; os.makedirs(OUT, exist_ok=True)

STAGE = {
    "New / Uncontacted": "new", "Contacted / Intake": "contacted",
    "Qualified / Interested": "qualified", "Site Visit / Showroom": "site_visit",
    "Proposal / Quote": "proposal", "Won": "won", "Unqualified": "unqualified",
    "Lost": "lost", "Duplicate": "duplicate", "No Answer": "no_answer",
}
SOURCE = {"FACEBOOK": "FACEBOOK", "WEBSITE": "WEBSITE", "LINE": "LINE", "CALL": "CALL", "OTHER": "OTHER"}
CHANNEL = {
    "Phone/LINE": "PHONE", "LINE": "LINE", "Messenger": "MESSENGER", "Facebook": "MESSENGER",
    "Website": "SITE_FORM", "Email": "EMAIL", "WhatsApp": "WHATSAPP", "Call": "PHONE",
}
DIRECTION = {"Outbound": "OUT", "Inbound": "IN"}
PRIORITY = {"High": "high", "Medium": "medium", "Low": "low", "Urgent": "urgent"}
TASK_TYPE = {"Follow-up": "โทรติดตาม", "Site Visit": "นัดดูโชว์รูม/ที่ดิน", "Proposal": "ส่งข้อเสนอ",
             "Call": "โทรติดตาม", "LINE": "โทรติดตาม"}
FU_STATUS = {"Open": "open", "Done": "done", "Cancelled": "cancelled"}

def s(v):
    if v is None: return None
    v = str(v).strip()
    return v if v and v.lower() not in ("", "-", "none", "n/a") else None

def d(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    t = s(v)
    if not t: return None
    try: return datetime.datetime.strptime(t[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except Exception: return None

def dt(v):
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(v, datetime.date): return v.strftime("%Y-%m-%dT00:00:00")
    t = s(v)
    if not t: return None
    try: return datetime.datetime.strptime(t[:19], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        try: return datetime.datetime.strptime(t[:10], "%Y-%m-%d").strftime("%Y-%m-%dT00:00:00")
        except Exception: return None

def num(v):
    if v is None: return None
    try: return float(v)
    except Exception: return None

def yn(v):
    t = s(v)
    if not t: return False
    return t.lower() in ("yes", "y", "true", "1", "ใช่")

def norm_phone(v):
    t = s(v)
    if not t: return None
    p = "".join(ch for ch in t if ch.isdigit())
    if not p: return None
    if p.startswith("66") and len(p) == 11: p = "0" + p[2:]
    if p.startswith("+66"): p = "0" + p[3:]
    if len(p) == 9 and p.startswith("8"): p = "0" + p
    return p if len(p) >= 9 else None

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# ---- CONTACTS ----
contacts, contact_phone_map = [], {}
for row in wb["Contacts"].iter_rows(values_only=True):
    h = ["Contact ID","Full Name","Primary Phone","Normalized Phone","Email","LINE ID","Facebook / Messenger Name",
         "Province / Location","Preferred Contact Time","First Source","First Lead Date","Latest Lead ID",
         "Client Interest / Project Type","Budget Range","Land Status","Timeline","CRM Stage","Contacted Yet?",
         "Last Contact Date","Last Contact Channel","Last Contact Outcome","Latest Conversation Summary",
         "Next Follow-up Date","Next Action","Owner / Salesperson","Priority","Do Not Contact?","Contact Notes",
         "All Linked Lead IDs","All Log Count"]
    r = dict(zip(h, row))
    if not s(r.get("Contact ID")): continue
    np = norm_phone(r.get("Primary Phone")) or norm_phone(r.get("Normalized Phone"))
    c = {
        "legacy_id": s(r.get("Contact ID")), "full_name": s(r.get("Full Name")) or "ไม่ระบุชื่อ",
        "primary_phone": s(r.get("Primary Phone")), "normalized_phone": np,
        "email": s(r.get("Email")), "line_id": s(r.get("LINE ID")), "fb_name": s(r.get("Facebook / Messenger Name")),
        "province": s(r.get("Province / Location")), "preferred_contact_time": s(r.get("Preferred Contact Time")),
        "first_source": SOURCE.get(s(r.get("First Source")), "OTHER") if s(r.get("First Source")) else None,
        "first_lead_date": dt(r.get("First Lead Date")), "latest_lead_id": s(r.get("Latest Lead ID")),
        "interest": s(r.get("Client Interest / Project Type")), "budget_range": s(r.get("Budget Range")),
        "land_status": s(r.get("Land Status")), "timeline": s(r.get("Timeline")),
        "crm_stage": STAGE.get(s(r.get("CRM Stage")), "new") if s(r.get("CRM Stage")) else "new",
        "contacted_yet": yn(r.get("Contacted Yet?")),
        "last_contact_date": dt(r.get("Last Contact Date")),
        "last_contact_channel": CHANNEL.get(s(r.get("Last Contact Channel")), "OTHER") if s(r.get("Last Contact Channel")) else None,
        "last_contact_outcome": s(r.get("Last Contact Outcome")),
        "latest_conversation_summary": s(r.get("Latest Conversation Summary")),
        "next_followup_date": dt(r.get("Next Follow-up Date")), "next_action": s(r.get("Next Action")),
        "owner": s(r.get("Owner / Salesperson")),
        "priority": PRIORITY.get(s(r.get("Priority")), "medium") if s(r.get("Priority")) else "medium",
        "do_not_contact": yn(r.get("Do Not Contact?")), "notes": s(r.get("Contact Notes")),
        "all_linked_lead_ids": [x for x in (str(r.get("All Linked Lead IDs") or "").split(",")) if x.strip()],
        "log_count": num(r.get("All Log Count")) or 0,
    }
    contacts.append(c)
    if np and c["legacy_id"]: contact_phone_map.setdefault(np, []).append(c["legacy_id"])

# ---- LEADS ----
leads = []
for row in wb["Leads"].iter_rows(values_only=True):
    h = ["Lead ID","Contact ID","Lead Date","Month Tab","Lead Status (Original)","Contact Status (Original)",
         "CRM Stage","Priority","Full Name","Phone Number","Email Address","Province / Location","Source",
         "Interest / ข้อมูล","Preferred Contact Time","Owner / Called By","Next Follow-up Date","Last Contact Date",
         "Last Contact Channel","Last Conversation Summary","Deal Value Estimate","Probability %","Next Action",
         "Source Sheet Row","Archived?","Meta Lead ID","Duplicate Check","Duplicate Match"]
    r = dict(zip(h, row))
    if not s(r.get("Lead ID")): continue
    leads.append({
        "legacy_id": s(r.get("Lead ID")), "legacy_contact_id": s(r.get("Contact ID")),
        "lead_date": dt(r.get("Lead Date")), "month_tab": s(r.get("Month Tab")),
        "original_statuses": s(r.get("Lead Status (Original)")),
        "crm_stage": STAGE.get(s(r.get("CRM Stage")), "new") if s(r.get("CRM Stage")) else "new",
        "priority": PRIORITY.get(s(r.get("Priority")), "medium") if s(r.get("Priority")) else "medium",
        "full_name": s(r.get("Full Name")) or "ไม่ระบุชื่อ", "phone": s(r.get("Phone Number")),
        "email": s(r.get("Email Address")), "province": s(r.get("Province / Location")),
        "source": SOURCE.get(s(r.get("Source")), "OTHER") if s(r.get("Source")) else "OTHER",
        "interest": s(r.get("Interest / ข้อมูล")), "preferred_contact_time": s(r.get("Preferred Contact Time")),
        "owner": s(r.get("Owner / Called By")), "next_followup_date": dt(r.get("Next Follow-up Date")),
        "last_contact_date": dt(r.get("Last Contact Date")),
        "last_contact_channel": CHANNEL.get(s(r.get("Last Contact Channel")), "OTHER") if s(r.get("Last Contact Channel")) else None,
        "last_contact_summary": s(r.get("Last Conversation Summary")),
        "deal_value": num(r.get("Deal Value Estimate")), "probability_pct": num(r.get("Probability %")),
        "next_action": s(r.get("Next Action")), "archived": yn(r.get("Archived?")),
        "meta_lead_id": s(r.get("Meta Lead ID")), "duplicate_check": s(r.get("Duplicate Check")),
    })

# ---- LOGS ----
logs = []
for row in wb["Conversation Logs"].iter_rows(values_only=True):
    h = ["Log ID","Lead ID","Contact ID","Date/Time","Channel","Direction","Team Member","Outcome",
         "Conversation Summary / Notes","Next Action","Next Follow-up Date","Attachment / Link"]
    r = dict(zip(h, row))
    if not s(r.get("Log ID")): continue
    logs.append({
        "legacy_id": s(r.get("Log ID")), "legacy_lead_id": s(r.get("Lead ID")), "legacy_contact_id": s(r.get("Contact ID")),
        "logged_at": dt(r.get("Date/Time")),
        "channel": CHANNEL.get(s(r.get("Channel")), "OTHER") if s(r.get("Channel")) else None,
        "direction": DIRECTION.get(s(r.get("Direction")), "IN") if s(r.get("Direction")) else "IN",
        "team_member": s(r.get("Team Member")), "outcome": s(r.get("Outcome")),
        "summary": s(r.get("Conversation Summary / Notes")), "next_action": s(r.get("Next Action")),
        "next_followup_date": dt(r.get("Next Follow-up Date")), "attachment_link": s(r.get("Attachment / Link")),
    })

# ---- FOLLOW UPS ----
followups = []
for row in wb["Follow Ups"].iter_rows(values_only=True):
    h = ["Task ID","Lead ID","Contact ID","Due Date","Due Time","Task Type","Owner","Priority","Status","Latest Note"]
    r = dict(zip(h, row))
    if not s(r.get("Task ID")): continue
    tt = s(r.get("Task Type"))
    followups.append({
        "legacy_id": s(r.get("Task ID")), "legacy_lead_id": s(r.get("Lead ID")), "legacy_contact_id": s(r.get("Contact ID")),
        "due_date": d(r.get("Due Date")), "due_time": s(r.get("Due Time")),
        "task_type": TASK_TYPE.get(tt, tt) if tt else "โทรติดตาม",
        "owner": s(r.get("Owner")), "priority": PRIORITY.get(s(r.get("Priority")), "medium") if s(r.get("Priority")) else "medium",
        "status": FU_STATUS.get(s(r.get("Status")), "open") if s(r.get("Status")) else "open",
        "latest_note": s(r.get("Latest Note")),
    })

out = {"contacts": contacts, "leads": leads, "logs": logs, "follow_ups": followups}
with open(f"{OUT}/payload.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False)
print(f"contacts={len(contacts)} leads={len(leads)} logs={len(logs)} followups={len(followups)}")
print(f"phone-collision contacts: {sum(1 for v in contact_phone_map.values() if len(v)>1)}")
print(f"payload bytes: {os.path.getsize(f'{OUT}/payload.json')}")
print("stage dist:", {k: sum(1 for l in leads if l['crm_stage']==k) for k in sorted(set(l['crm_stage'] for l in leads))})
